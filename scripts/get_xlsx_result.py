import json
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils import get_column_letter
from typing import List, Tuple
from pathlib import Path

def setup_excel_styles() -> Border:
    """Create and return Excel cell border styles"""
    return Border(left=Side(style='thin'), 
                 right=Side(style='thin'),
                 top=Side(style='thin'),
                 bottom=Side(style='thin'))

def write_excel_cell(ws, row: int, col: int, value, border: Border, alignment: Alignment = None) -> None:
    """Write a value to Excel cell with specified styles"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = border
    if alignment:
        cell.alignment = alignment

def process_metrics_data(metrics_path: Path) -> Tuple[List, List]:
    """Process metrics JSON data and return formatted data"""
    try:
        with open(metrics_path) as f:
            metrics = json.load(f)
        
        data = []
        mean_values = []
        for obj, values in metrics.items():
            if isinstance(values, dict):
                data.append([obj] + list(values.values()))
            else:
                mean_values.append(values)
        data.append(['Mean', *mean_values])
        
        return data, mean_values
    except Exception as e:
        print(f"Error processing metrics file {metrics_path}: {str(e)}")
        return [], []

def create_summary_excel(results_root: str, setting_tag: str, n_shots: List[int], a_shots: List[int], num_seeds: int) -> None:
    """Create summary Excel file for all datasets"""
    try:
        results_path = Path(results_root) / setting_tag
        excel_metrics_path = results_path / 'excel_metrics'
        excel_metrics_path.mkdir(parents=True, exist_ok=True)

        # Get dataset list
        dataset = sorted([d for d in results_path.iterdir() if d.is_dir()])[:-1]
        if not dataset:
            raise ValueError("No datasets found in the specified directory")

        # Define title structure
        titles = [
            [d.name for d in dataset],  # Dataset names
            ['seg_AUPRO', 'seg_AUROC', 'seg_F1', 'cls_AUROC', 'cls_AP', 'cls_F1'],
            n_shots,
            a_shots,
        ]

        for seed in range(num_seeds):
            all_data = np.zeros((len(dataset), 6, len(n_shots), len(a_shots)))
            metrics_exist = False

            # Process each dataset
            for d_i, d in enumerate(dataset):
                # Create a new Excel writer for each dataset
                xlsx_path = excel_metrics_path / f'{setting_tag}_{d.name}_metrics_seed={seed}.xlsx'
                with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                    for n_i, n in enumerate(n_shots):
                        for a_i, a in enumerate(a_shots):
                            shots_setting = f'{n}-n_shot_{a}-a_shot'
                            metrics_path = d / shots_setting / f'metrics_seed={seed}.json'
                            
                            if metrics_path.exists():
                                metrics_exist = True
                                data, mean_values = process_metrics_data(metrics_path)
                                
                                if data and mean_values:
                                    # Create sheet name
                                    sheet_name = f'{n}n_{a}a'
                                    # Convert data to DataFrame and save to Excel
                                    df = pd.DataFrame(data, columns=['objects', *titles[1]])
                                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                                    all_data[d_i, :, n_i, a_i] = mean_values

            # Create summary Excel if metrics exist
            if metrics_exist:
                create_summary_worksheet(results_path, titles, all_data, seed)

    except Exception as e:
        print(f"Error creating summary Excel: {str(e)}")


def create_summary_worksheet(results_path: Path, titles: List, all_data: np.ndarray, seed: int) -> None:
    """Create summary worksheet with all metrics"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        border = setup_excel_styles()
        center_align = Alignment(horizontal='center')
        center_vertical_align = Alignment(horizontal='center', vertical='center')
        
        rows_per_block = len(titles[2])
        cols_per_subblock = len(titles[3])
        current_row = 3
        
        # Write headers
        for j, dim2_title in enumerate(titles[1]):
            start_col = 3 + j * cols_per_subblock
            
            # Write second dimension title
            write_excel_cell(ws, 1, start_col, dim2_title, border, center_align)
            merge_range = f'{get_column_letter(start_col)}{1}:{get_column_letter(start_col+cols_per_subblock-1)}{1}'
            ws.merge_cells(merge_range)
            
            # Write fourth dimension title
            for k, dim4_title in enumerate(titles[3]):
                write_excel_cell(ws, 2, start_col + k, dim4_title, border, center_align)
        
        # Write data blocks
        for d1_i, d1_title in enumerate(titles[0]):
            # Write first dimension title
            write_excel_cell(ws, current_row, 1, d1_title, border, center_vertical_align)
            ws.merge_cells(f'A{current_row}:A{current_row+rows_per_block-1}')
            
            # Write third dimension title
            for i, dim3_title in enumerate(titles[2]):
                write_excel_cell(ws, current_row + i, 2, dim3_title, border, center_align)
            
            # Write metrics data
            for j in range(len(titles[1])):
                start_col = 3 + j * cols_per_subblock
                for i in range(rows_per_block):
                    for k in range(cols_per_subblock):
                        value = round(all_data[d1_i, j, i, k], 16)  # Reduced decimal places for better readability
                        write_excel_cell(ws, current_row+i, start_col+k, value, border, center_align)
            
            current_row += rows_per_block + 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        for col in range(3, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        wb.save(results_path / 'excel_metrics' / f'all_metrics_seed={seed}.xlsx')

    except Exception as e:
        print(f"Error creating summary worksheet: {str(e)}")

if __name__ == "__main__":
    results_root = 'results/dinov2_vits14_448'
    setting_tag = 'default'
    n_shots = [1, 2, 4]
    a_shots = [1]
    num_seeds = 3

    create_summary_excel(results_root, setting_tag, n_shots, a_shots, num_seeds)
